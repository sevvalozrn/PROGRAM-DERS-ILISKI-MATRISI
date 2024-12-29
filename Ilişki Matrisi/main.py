import pyodbc
from openpyxl import Workbook
from openpyxl.comments import Comment


def get_connection(database=None):
    connection_string = (
        "Driver={ODBC Driver 17 for SQL Server};"
        "Server=ASUS\\SQLEXPRESS;"
        f"Database={database or 'master'};"
        "Trusted_Connection=yes;"
    )
    return pyodbc.connect(connection_string)


def check_database():
    conn = get_connection("master")
    conn.autocommit = True

    cursor = conn.cursor()

    cursor.execute(f"SELECT database_id FROM sys.databases WHERE Name = 'RelationMatrix'")
    database_exists = cursor.fetchone()

    if database_exists:
        print("There is already a database called RelationMatrix")
    else:
        cursor.execute("CREATE DATABASE RelationMatrix")
        print("RelationMatrix database created")

    conn.close()


def check_tables():
    conn = get_connection("RelationMatrix")
    conn.autocommit = True
    cursor = conn.cursor()

    tables = ["CourseOutcomes", "ProgramOutcomes", "ProgramCourseRelations", "EvaluationCriteria",
              "CourseEvaluationRelations"]
    for table in tables:
        cursor.execute(f"SELECT COUNT(*) FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_NAME = '{table}'")
        table_exists = cursor.fetchone()[0]
        if table_exists:
            print(f"{table} table already exists.")
        else:
            if table == 'CourseOutcomes':
                cursor.execute('''
                    CREATE TABLE CourseOutcomes (
                        id INT PRIMARY KEY IDENTITY(1,1),
                        data TEXT NOT NULL
                    );
                ''')
            elif table == 'ProgramOutcomes':
                cursor.execute('''
                    CREATE TABLE ProgramOutcomes (
                        id INT PRIMARY KEY IDENTITY(1,1),
                        data TEXT NOT NULL
                    );
                ''')
            elif table == 'ProgramCourseRelations':
                cursor.execute('''
                    CREATE TABLE ProgramCourseRelations (
                        ProgramOutcomeID INT NOT NULL,
                        CourseOutcomeID INT NOT NULL,
                        RelationValue FLOAT NOT NULL CHECK (RelationValue BETWEEN 0 AND 1),
                        PRIMARY KEY (ProgramOutcomeID, CourseOutcomeID),
                        FOREIGN KEY (ProgramOutcomeID) REFERENCES ProgramOutcomes(id),
                        FOREIGN KEY (CourseOutcomeID) REFERENCES CourseOutcomes(id)
                    );
                ''')
            elif table == 'EvaluationCriteria':
                cursor.execute('''
                    CREATE TABLE EvaluationCriteria (
                        Criteria VARCHAR(25) PRIMARY KEY,
                        Weight INT NOT NULL
                    );
                ''')
            elif table == 'CourseEvaluationRelations':
                cursor.execute('''
                    CREATE TABLE CourseEvaluationRelations (
                        CourseOutcomeID INT NOT NULL,
                        Criteria VARCHAR(25),
                        RelationValue INT NOT NULL,
                        PRIMARY KEY (CourseOutcomeID, Criteria),
                        FOREIGN KEY (CourseOutcomeID) REFERENCES CourseOutcomes(id),
                        FOREIGN KEY (Criteria) REFERENCES EvaluationCriteria(Criteria),
                    );
                ''')
            print(f"{table} table created.")

    conn.close()


def insert_data_into_table(table_name, data):
    conn = get_connection("RelationMatrix")
    conn.autocommit = True
    cursor = conn.cursor()

    if table_name == 'CourseOutcomes':
        cursor.execute("INSERT INTO CourseOutcomes (data) VALUES (?)", data)
    elif table_name == 'ProgramOutcomes':
        cursor.execute("INSERT INTO ProgramOutcomes (data) VALUES (?)", data)

    conn.close()


def fetch_relations():
    conn = get_connection("RelationMatrix")
    cursor = conn.cursor()

    query = """
    SELECT ProgramOutcomeID, CourseOutcomeID, RelationValue
    FROM ProgramCourseRelations;
    """
    cursor.execute(query)

    relations = cursor.fetchall()

    cursor.close()
    conn.close()

    return relations


def fetch_evaluation_relations():
    conn = get_connection("RelationMatrix")
    cursor = conn.cursor()

    query = """
    SELECT CourseOutcomeID, Criteria ,RelationValue
    FROM CourseEvaluationRelations;
    """
    cursor.execute(query)

    relations = cursor.fetchall()

    cursor.close()
    conn.close()

    return relations


def fetch_table_data(table_name):
    conn = get_connection("RelationMatrix")
    cursor = conn.cursor()

    query = f"SELECT id, data FROM {table_name};"
    cursor.execute(query)

    data = cursor.fetchall()

    cursor.close()
    conn.close()

    return data


def fetch_evaluation_data():
    conn = get_connection("RelationMatrix")
    cursor = conn.cursor()

    query = "SELECT Criteria, Weight FROM EvaluationCriteria;"
    cursor.execute(query)
    data = cursor.fetchall()

    cursor.close()
    conn.close()

    return data


def fetch_student_data():
    conn = get_connection("RelationMatrix")
    cursor = conn.cursor()

    query = "SELECT Student FROM Students;"
    cursor.execute(query)
    students = cursor.fetchall()

    cursor.close()
    conn.close()

    return students


def fetch_success_rate():
    conn = get_connection("RelationMatrix")
    cursor = conn.cursor()

    query = "SELECT student_id, success_rate FROM Table4;"
    cursor.execute(query)
    success_rates = cursor.fetchall()

    cursor.close()
    conn.close()

    return success_rates


def create_table1():
    workbook = Workbook()
    sheet = workbook.active

    program_outcomes = fetch_table_data("ProgramOutcomes")
    course_outcomes = fetch_table_data("CourseOutcomes")

    program_row_count = len(program_outcomes)
    course_row_count = len(course_outcomes)

    sheet.merge_cells('A1:B1')
    sheet['A1'] = "Table 1"

    sheet.merge_cells('A2:B2')
    sheet['A2'] = "Program Outcomes"

    sheet.merge_cells(start_row=1, start_column=3, end_row=1, end_column=course_row_count + 2)
    sheet['C1'] = "Course Outcomes"

    for i, (program_id, program_text) in enumerate(program_outcomes, start=1):
        sheet.merge_cells(f"A{i + 2}:B{i + 2}")
        cell = sheet[f"A{i + 2}"]
        cell.value = program_id
        comment = Comment(program_text, "Database")
        cell.comment = comment

    course_outcomes = fetch_table_data("CourseOutcomes")
    for j, (course_id, course_text) in enumerate(course_outcomes, start=1):
        c = sheet.cell(row=2, column=j + 2)
        c.value = course_id
        comment = Comment(course_text, "Database")
        c.comment = comment

    relations = fetch_relations()
    for relation in relations:
        program_outcome_id, course_outcome_id, relation_value = relation
        row = program_outcome_id + 2
        col = course_outcome_id + 2
        sheet.cell(row=row, column=col, value=relation_value)

    for i in range(3, program_row_count + 3):
        total = sum(sheet.cell(row=i, column=j).value or 0 for j in range(3, course_row_count + 3))
        result = round((total / course_row_count), 2)
        sheet.cell(row=i, column=course_row_count + 3, value=result)

    sheet.cell(row=2, column=course_row_count + 3, value="Rel Value")

    workbook.save(filename="table1.xlsx")


def create_table2():
    workbook = Workbook()
    sheet = workbook.active

    course_outcomes = fetch_table_data("CourseOutcomes")

    sheet.merge_cells('A1:B1')
    sheet['A1'] = "Table 2"

    sheet.merge_cells('A2:B2')
    sheet['A2'] = "Course Outcomes"

    for i, (program_id, program_text) in enumerate(course_outcomes, start=1):
        sheet.merge_cells(f"A{i + 2}:B{i + 2}")
        cell = sheet[f"A{i + 2}"]
        cell.value = program_id
        comment = Comment(program_text, "Database")
        cell.comment = comment

    criteria_data = fetch_evaluation_data()
    for index, (criteria, weight) in enumerate(criteria_data, start=3):
        sheet.cell(row=1, column=index, value=weight)
        sheet.cell(row=2, column=index, value=criteria)

        relations = fetch_evaluation_relations()
    for relation in relations:
        course_outcome_id, criteria, relation_value = relation

        # CourseOutcomeID'ye göre satır bulunması
        row = None
        for i, (course_id, _) in enumerate(course_outcomes, start=3):
            if course_outcome_id == course_id:
                row = i
                break

        # Criteria'ya göre sütun bulunması
        col = None
        for j, (criterion, _) in enumerate(criteria_data, start=3):
            if criteria == criterion:
                col = j
                break

        # Değeri tabloya yazdırma
        if row and col:
            sheet.cell(row=row, column=col, value=relation_value)

    # Her satır için toplam hesaplama
    total_col = len(criteria_data) + 3
    sheet.cell(row=2, column=total_col, value="Total")
    for row_idx in range(3, len(course_outcomes) + 3):
        total = sum(sheet.cell(row=row_idx, column=col_idx).value or 0 for col_idx in range(3, total_col))
        sheet.cell(row=row_idx, column=total_col, value=total)

    workbook.save(filename="table2.xlsx")


def create_table3():
    workbook = Workbook()
    sheet = workbook.active

    sheet.merge_cells('A1:B1')
    sheet['A1'] = "Table 3"
    sheet['C1'] = "Weighted Evaluation"

    sheet.merge_cells('A2:B2')
    sheet['A2'] = "Course Outcomes"

    evaluation_criteria = fetch_evaluation_data()
    criteria_weights = {criteria: weight for criteria, weight in evaluation_criteria}
    criteria_headers = list(criteria_weights.keys())

    for col, criteria in enumerate(criteria_headers, start=3):
        sheet.cell(row=2, column=col, value=criteria)

    total_col = len(criteria_headers) + 3
    sheet.cell(row=2, column=total_col, value="Total")

    course_evaluation_relations = fetch_evaluation_relations()
    course_outcomes = fetch_table_data("CourseOutcomes")

    weighted_data = {}
    for course_outcome_id, criteria, relation_value in course_evaluation_relations:
        weight = criteria_weights.get(criteria, 0)
        weighted_value = (relation_value * weight) / 100

        if course_outcome_id not in weighted_data:
            weighted_data[course_outcome_id] = {}

        weighted_data[course_outcome_id][criteria] = weighted_value

    for row_idx, (course_outcome_id, program_text) in enumerate(course_outcomes, start=3):
        sheet.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
        cell = sheet.cell(row=row_idx, column=1, value=course_outcome_id)

        comment = Comment(program_text, "Database")
        cell.comment = comment

        total = 0
        for col_idx, criteria in enumerate(criteria_headers, start=3):
            value = weighted_data.get(course_outcome_id, {}).get(criteria, 0)
            sheet.cell(row=row_idx, column=col_idx, value=value)
            total += value

        sheet.cell(row=row_idx, column=total_col, value=total)

    workbook.save("table3.xlsx")


def save_table3_to_database():
    evaluation_criteria = fetch_evaluation_data()
    criteria_weights = {criteria: weight for criteria, weight in evaluation_criteria}
    course_evaluation_relations = fetch_evaluation_relations()
    course_outcomes = fetch_table_data("CourseOutcomes")

    weighted_data = {}
    for course_outcome_id, criteria, relation_value in course_evaluation_relations:
        weight = criteria_weights.get(criteria, 0)
        weighted_value = (relation_value * weight) / 100

        if course_outcome_id not in weighted_data:
            weighted_data[course_outcome_id] = {}

        weighted_data[course_outcome_id][criteria] = weighted_value

    conn = get_connection("RelationMatrix")
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_NAME = 'Table3';")
    existing_table = cursor.fetchone()

    if existing_table:
        cursor.execute("DROP TABLE Table3;")

    criteria_columns = ", ".join(f'"{criteria}" FLOAT' for criteria in criteria_weights.keys())
    cursor.execute(f"""CREATE TABLE Table3 (id INT IDENTITY(1,1) PRIMARY KEY,course_outcome_id 
                   INT NOT NULL,total_score FLOAT NOT NULL,{criteria_columns},)""")

    for course_outcome_id, program_text in course_outcomes:
        total_score = 0
        criteria_values = []

        for criteria in criteria_weights.keys():
            value = weighted_data.get(course_outcome_id, {}).get(criteria, 0)
            criteria_values.append(value)
            total_score += value

        columns = ", ".join(f'"{criteria}"' for criteria in criteria_weights.keys())
        placeholders = ", ".join("?" for _ in criteria_values)

        cursor.execute(f"""INSERT INTO Table3 (course_outcome_id, total_score, {columns})
            VALUES (?, ?, {placeholders})""", (course_outcome_id, total_score, *criteria_values))

    conn.commit()
    conn.close()


def create_notes():
    workbook = Workbook()
    sheet = workbook.active

    conn = get_connection("RelationMatrix")
    cursor = conn.cursor()

    sheet.merge_cells('A1:B1')
    sheet['A1'] = "Table Note"
    sheet['C1'] = "Notes"

    cursor.execute("SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS WHERE TABLE_NAME = 'Students';")
    columns = [row[0] for row in cursor.fetchall()]

    if not columns:
        print("Students table is empty or does not exist.")
        return

    cursor.execute("SELECT * FROM Students;")
    rows = cursor.fetchall()

    if not rows:
        print("No data found in Students table.")
        return

    cursor.execute("SELECT Criteria, Weight FROM EvaluationCriteria;")
    criteria_weights = {row[0]: row[1] for row in cursor.fetchall()}

    for col_idx, column_name in enumerate(columns, start=1):
        sheet.cell(row=2, column=col_idx + 1, value=column_name)

    sheet.cell(row=2, column=len(columns) + 2, value="Average")

    for row_idx, row in enumerate(rows, start=3):
        total_score = 0
        weight_sum = 0

        sheet.merge_cells(f'A{row_idx}:B{row_idx}')
        sheet[f'A{row_idx}'] = row[0]

        for col_idx, value in enumerate(row[1:], start=2):
            sheet.cell(row=row_idx, column=col_idx + 1, value=value)

            criterion = columns[col_idx - 1]
            weight = criteria_weights.get(criterion, 0)
            total_score += value * weight
            weight_sum += weight

        average = total_score / weight_sum if weight_sum > 0 else 0
        sheet.cell(row=row_idx, column=len(columns) + 2, value=round(average, 2))

    workbook.save("notlar.xlsx")
    conn.close()


def create_table4():
    conn = get_connection("RelationMatrix")
    cursor = conn.cursor()

    cursor.execute("SELECT Criteria, Weight FROM EvaluationCriteria;")
    evaluation_weights = {row[0]: row[1] for row in cursor.fetchall()}

    cursor.execute("SELECT * FROM Students;")
    student_columns = [column[0] for column in cursor.description]
    student_data = cursor.fetchall()

    cursor.execute("SELECT id, data FROM CourseOutcomes;")
    course_outcomes = cursor.fetchall()

    cursor.execute("SELECT CourseOutcomeID, Criteria, RelationValue FROM CourseEvaluationRelations;")
    course_evaluation_relations = cursor.fetchall()

    conn.close()

    weighted_values = {}
    for outcome_id, criteria, relation_value in course_evaluation_relations:
        weight = evaluation_weights.get(criteria, 0)
        if outcome_id not in weighted_values:
            weighted_values[outcome_id] = {}
        weighted_values[outcome_id][criteria] = relation_value * (weight / 100)

    workbook = Workbook()
    sheet_created = False

    for student in student_data:
        if not sheet_created:
            sheet = workbook.active
            sheet.title = f"Student {student[0]}"
            sheet_created = True
        else:
            sheet = workbook.create_sheet(title=f"Student {student[0]}")

        sheet.merge_cells("A1:B1")
        sheet["A1"] = "Table 4"

        sheet.merge_cells("C1:D1")
        sheet["C1"] = f"Student {student[0]}"

        sheet.merge_cells("A2:B2")
        sheet["A2"] = "Program Outcomes"

        headers = [*evaluation_weights.keys(), "Total", "Max", "% Success"]
        for col_index, header in enumerate(headers, start=3):
            sheet.cell(row=2, column=col_index, value=header)

        for outcome_id, outcome_text in course_outcomes:
            current_row = sheet.max_row + 1
            sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
            cell = sheet.cell(row=current_row, column=1, value=f"{outcome_text:>35}")
            row = []
            total = 0

            for criteria in evaluation_weights.keys():
                score = student[student_columns.index(criteria)]
                weight_value = weighted_values.get(outcome_id, {}).get(criteria, 0)
                weighted_score = score * weight_value
                row.append(weighted_score)
                total += weighted_score

            max_score = sum(weighted_values.get(outcome_id, {}).values()) * 100
            success_rate = (total / max_score * 100) if max_score > 0 else 0

            row.extend([total, max_score, round(success_rate, 1)])

            for col_index, value in enumerate(row, start=3):
                sheet.cell(row=current_row, column=col_index, value=value)

            cell = sheet.cell(row=current_row, column=1)
            cell.comment = Comment(outcome_text, "Database")

    workbook.save("table4.xlsx")


def save_table4_to_database():
    conn = get_connection("RelationMatrix")
    cursor = conn.cursor()

    cursor.execute("SELECT Criteria, Weight FROM EvaluationCriteria;")
    evaluation_weights = {row[0]: row[1] for row in cursor.fetchall()}

    cursor.execute("SELECT * FROM Students;")
    student_columns = [column[0] for column in cursor.description]
    student_data = cursor.fetchall()

    cursor.execute("SELECT id, data FROM CourseOutcomes;")
    course_outcomes = cursor.fetchall()
    cursor.execute("SELECT CourseOutcomeID, Criteria, RelationValue FROM CourseEvaluationRelations;")
    course_evaluation_relations = cursor.fetchall()
    conn.close()

    weighted_values = {}
    for outcome_id, criteria, relation_value in course_evaluation_relations:
        weight = evaluation_weights.get(criteria, 0)
        if outcome_id not in weighted_values:
            weighted_values[outcome_id] = {}
        weighted_values[outcome_id][criteria] = relation_value * (weight / 100)

    conn = get_connection("RelationMatrix")
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_NAME = 'Table4';")
    existing_table = cursor.fetchone()

    if existing_table:
        cursor.execute("DROP TABLE Table4;")

    criteria_columns = ", ".join(f'"{criteria}" FLOAT' for criteria in evaluation_weights.keys())
    cursor.execute(f"""
        CREATE TABLE Table4 (
            id INT IDENTITY(1,1) PRIMARY KEY,
            student_id INT NOT NULL,
            course_outcome_id INT NOT NULL,
            total_score FLOAT NOT NULL,
            {criteria_columns},
            max_score FLOAT,
            success_rate FLOAT
        );
    """)

    for student in student_data:
        student_id = student[0]

        for outcome_id, outcome_text in course_outcomes:
            total_score = 0
            criteria_values = []

            for criteria in evaluation_weights.keys():
                score = student[student_columns.index(criteria)]
                weight_value = weighted_values.get(outcome_id, {}).get(criteria, 0)
                weighted_score = score * weight_value
                criteria_values.append(weighted_score)
                total_score += weighted_score

            max_score = sum(weighted_values.get(outcome_id, {}).values()) * 100
            success_rate = (total_score / max_score * 100) if max_score > 0 else 0

            columns = ", ".join(f'"{criteria}"' for criteria in evaluation_weights.keys())
            placeholders = ", ".join("?" for _ in criteria_values)

            cursor.execute(f"""
                INSERT INTO Table4 (student_id, course_outcome_id, total_score, {columns}, max_score, success_rate)
                VALUES (?, ?, ?, {placeholders}, ?, ?)
            """, (student_id, outcome_id, total_score, *criteria_values, max_score, round(success_rate, 2)))

    conn.commit()
    conn.close()


def create_table5():
    workbook = Workbook()
    students = fetch_student_data()
    course_outcomes = fetch_table_data("CourseOutcomes")
    program_outcomes = fetch_table_data("ProgramOutcomes")
    relations = fetch_relations()
    success_rates = fetch_success_rate()

    for student in students:
        student_id = student[0]

        sheet = workbook.create_sheet(title=f"Student {student_id}")
        row = 1

        sheet.merge_cells("A1:B1")
        sheet["A1"] = f"Student {student_id}"
        row += 1

        sheet.cell(row=row, column=1, value="Table 5")
        for idx, outcome in enumerate(course_outcomes, start=2):
            sheet.cell(row=row, column=idx, value=outcome[1])
        row += 1

        sheet.cell(row=row, column=1, value="ProgramOutcomes")
        success_rate_for_student = [
            sr[1] for sr in success_rates if sr[0] == student_id
        ]
        for idx, success_rate in enumerate(success_rate_for_student, start=2):
            formatted_rate = round(success_rate, 1)
            sheet.cell(row=row, column=idx, value=formatted_rate)
        sheet.cell(row=row, column=idx + 1, value="Success Rate")
        row += 1

        for program_outcome in program_outcomes:
            row_values = []
            relation_values = []

            for course_outcome, success_rate in zip(course_outcomes, success_rate_for_student):
                relation_value = next(
                    (relation[2] for relation in relations if
                     relation[0] == program_outcome[0] and relation[1] == course_outcome[0]),
                    0
                )
                relation_values.append(relation_value)

            row_values = [
                relation * success_rate
                for relation, success_rate in zip(relation_values, success_rate_for_student)
            ]

            total_success = sum(row_values)
            num_course_outcomes = len(course_outcomes)
            avg_success = total_success / num_course_outcomes if num_course_outcomes else 0
            avg_relation_value = sum(relation_values) / len(relation_values) if relation_values else 0
            ratio = avg_success / avg_relation_value if avg_relation_value else 0

            program_outcome_id = program_outcome[0]
            program_outcome_data = program_outcome[1]
            cell = sheet.cell(row=row, column=1, value=program_outcome_id)

            comment_text = f"{program_outcome_data}"
            cell.comment = Comment(comment_text, "Generated")

            for idx, value in enumerate(row_values, start=2):
                sheet.cell(row=row, column=idx, value=round(value, 1))
            sheet.cell(row=row, column=idx + 1, value=round(ratio, 1))
            row += 1

    del workbook['Sheet']
    workbook.save("table5.xlsx")


def get_input_and_insert_relations():
    print("Enter 'q' to quit.")
    while True:
        program_outcome_id = input("Enter Program Outcome ID: ").strip()
        if program_outcome_id == 'q':
            print("Exiting the program.")
            break

        course_outcome_id = input("Enter Course Outcome ID: ").strip()
        if course_outcome_id == 'q':
            print("Exiting the program.")
            break

        relation_value = input("Enter Relation Value (0-1): ").strip()
        if relation_value == 'q':
            print("Exiting the program.")
            break

        try:
            relation_value = float(relation_value)
            if 0 <= relation_value <= 1:
                insert_relation_value(program_outcome_id, course_outcome_id, relation_value)
            else:
                print("Please enter a relation value between 0 and 1.")
        except ValueError:
            print("Invalid input for relation value. Please enter a valid number between 0 and 1.")


def get_input_and_insert_evaluation_relations():
    print("Enter 'q' to quit.")
    while True:
        course_outcome_id = input("Enter Course Outcome ID: ").strip()
        if course_outcome_id == 'q':
            print("Exiting the program.")
            break

        evaluation_criteria = input("Enter evaluation criteria: ").strip()
        if evaluation_criteria == 'q':
            print("Exiting the program.")
            break

        relation_value = input("Enter Relation Value (0/1): ").strip()
        if relation_value == 'q':
            print("Exiting the program.")
            break

        try:
            relation_value = int(relation_value)
            if (relation_value == 0 or relation_value == 1):
                insert_evaluation_relation_value(course_outcome_id, evaluation_criteria, relation_value)
            else:
                print("Please enter a relation value of 0 or 1 .")
        except ValueError:
            print("Invalid input for relation value. Please enter 0 or 1.")


def insert_relation_value(program_outcome_id, course_outcome_id, relation_value):
    conn = get_connection("RelationMatrix")
    conn.autocommit = True
    cursor = conn.cursor()

    cursor.execute('''
        INSERT INTO ProgramCourseRelations (ProgramOutcomeID, CourseOutcomeID, RelationValue)
        VALUES (?, ?, ?);
    ''', program_outcome_id, course_outcome_id, relation_value)

    print(
        f"Relation between ProgramOutcome {program_outcome_id} and CourseOutcome {course_outcome_id} has been inserted.")

    conn.close()


def insert_evaluation_relation_value(course_outcome_id, criteria, relation_value):
    conn = get_connection("RelationMatrix")
    conn.autocommit = True
    cursor = conn.cursor()

    cursor.execute('''
        INSERT INTO CourseEvaluationRelations (CourseOutcomeID, Criteria, RelationValue)
        VALUES (?, ?, ?);
    ''', course_outcome_id, criteria, relation_value)

    print(f"Relation between CourseOutcome {course_outcome_id} and EvaluationCriteria {criteria} has been inserted.")

    conn.close()


def get_input_and_insert_table(table_name):
    print("Enter 'q' to quit.")
    while True:
        data = input(f"Enter data for {table_name}: ").strip()
        if data == 'q':
            print(f"Exiting {table_name} input.")
            break

        insert_data_into_table(table_name, data)
        print(f"Data has been inserted into {table_name}.")


def get_evaluation_criteria_and_insert():
    print("Enter evaluation criteria and their weights. The total weight must be 100.")
    criteria_data = []
    total_weight = 0

    while True:
        criterion = input("Enter Criterion (or 'q' to quit): ").strip()
        if criterion.lower() == 'q':
            break

        weight = input(f"Enter Weight for {criterion}: ").strip()
        try:
            weight = int(weight)
            if weight < 0:
                print("Weight must be a positive integer.")
                continue
        except ValueError:
            print("Invalid weight. Please enter a positive integer.")
            continue

        total_weight += weight
        if total_weight > 100:
            print(f"Total weight exceeded 100 (current total: {total_weight}). Adjust your inputs.")
            total_weight -= weight
            continue

        criteria_data.append((criterion, weight))

        if total_weight == 100:
            break

        print(f"Current total weight: {total_weight}. You need {100 - total_weight} more.")

    if total_weight < 100:
        print(f"Total weight is {total_weight}, which is less than 100. Please try again.")
        return

    conn = get_connection("RelationMatrix")
    conn.autocommit = True
    cursor = conn.cursor()

    for criterion, weight in criteria_data:
        cursor.execute('''
            INSERT INTO EvaluationCriteria (Criteria, Weight)
            VALUES (?, ?);
        ''', criterion, weight)

    print("Evaluation criteria have been successfully inserted into the database.")
    conn.close()


is_table_created = False


def create_students_table():
    global is_table_created

    if is_table_created:
        return

    conn = get_connection("RelationMatrix")
    cursor = conn.cursor()

    cursor.execute("SELECT Criteria FROM EvaluationCriteria;")
    criteria = [row[0] for row in cursor.fetchall()]

    if not criteria:
        print("No evaluation criteria found. Please add criteria first.")
        conn.close()
        return

    cursor.execute("SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_NAME = 'Students';")
    existing_table = cursor.fetchone()

    if existing_table:
        print("Students table already exists.")
        cursor.execute("DROP TABLE Students;")

    columns = ["Student INT PRIMARY KEY"]
    for criterion in criteria:
        columns.append(f'[{criterion}] FLOAT')

    create_table_query = f"CREATE TABLE Students ({', '.join(columns)});"
    cursor.execute(create_table_query)
    conn.commit()

    conn.close()
    is_table_created = True


def add_student():
    conn = get_connection("RelationMatrix")
    cursor = conn.cursor()

    criteria_query = "SELECT Criteria FROM EvaluationCriteria;"
    cursor.execute(criteria_query)
    criteria = [row[0] for row in cursor.fetchall()]

    if not criteria:
        print("No evaluation criteria found. Please add criteria first.")
        conn.close()
        return

    while True:
        student_number = input("Enter Student Number (or 'q' to quit): ").strip()
        if student_number.lower() == 'q':
            print("Exiting student data entry.")
            break

        if not student_number.isdigit():
            print("Invalid Student Number. Please enter a valid number.")
            continue

        student_data = [int(student_number)]
        for criterion in criteria:
            while True:
                score = input(f"Enter score for {criterion} (0-100): ").strip()
                if not score.isdigit() and score.lower() != 'q':
                    print("Invalid input. Please enter a valid numeric score.")
                    continue
                if score.lower() == 'q':
                    print("Exiting student data entry.")
                    conn.close()
                    return
                score = float(score)
                if 0 <= score <= 100:
                    student_data.append(score)
                    break
                else:
                    print("Please enter a valid score between 0 and 100.")

        columns = ["Student"] + criteria
        placeholders = ", ".join(["?"] * len(columns))
        insert_query = f"INSERT INTO Students ({', '.join(columns)}) VALUES ({placeholders});"

        cursor.execute(insert_query, student_data)
        conn.commit()
        print("Student data has been successfully added.")
    conn.close()


def menu():
    while True:
        print("\nSelect an action:")
        print("For TABLE 1:")
        print("1. Add to Program Outcomes")
        print("2. Add to Course Outcomes")
        print("3. Add Relations")
        print("\nFor TABLE 2:")
        print("4. Add Evaluation Criteria")
        print("5. Add CourseOutcome-Criteria Relations")
        print("\nFor TABLE 4:")
        print("6. Add Student")
        print("\n7. Exit")

        choice = input("Enter your choice (1-7): ").strip()

        if choice == '1':
            get_input_and_insert_table('ProgramOutcomes')
        elif choice == '2':
            get_input_and_insert_table('CourseOutcomes')
        elif choice == '3':
            get_input_and_insert_relations()
        elif choice == '4':
            conn = get_connection("RelationMatrix")
            conn.autocommit = True
            cursor = conn.cursor()
            cursor.execute("DELETE FROM CourseEvaluationRelations")
            cursor.execute("DELETE FROM EvaluationCriteria")
            conn.close()
            get_evaluation_criteria_and_insert()
        elif choice == '5':
            get_input_and_insert_evaluation_relations()
        elif choice == '6':
            create_students_table()
            add_student()
        elif choice == '7':
            print("Exiting the program.")
            break
        else:
            print("Invalid choice. Please enter a number between 1 and 7.")


check_database()
check_tables()
menu()
create_table1()
create_table2()
create_table3()
save_table3_to_database()
create_notes()
create_table4()
save_table4_to_database()
create_table5()
